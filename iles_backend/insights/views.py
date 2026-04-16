from __future__ import annotations

import csv
from collections import Counter
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO

from django.http import HttpResponse
from django.utils import timezone
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.models import User
from accounts.permissions import IsAdmin
from evaluations.models import FinalInternshipScore, PlacementEvaluation
from logbook.models import WeeklyLog
from placements.models import Placement

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle


def _parse_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _apply_date_window(queryset, date_from=None, date_to=None, field_name='created_at'):
    if date_from:
        queryset = queryset.filter(**{f'{field_name}__date__gte': date_from})
    if date_to:
        queryset = queryset.filter(**{f'{field_name}__date__lte': date_to})
    return queryset


def _decimal_average(values):
    if not values:
        return Decimal('0.00')
    total = sum((Decimal(str(value)) for value in values), Decimal('0.00'))
    return (total / Decimal(len(values))).quantize(Decimal('0.01'))


def _top_items(counter, limit=5, total=None):
    items = []
    for label, count in counter.most_common(limit):
        share = round((count / total) * 100, 1) if total else 0
        items.append({'label': label, 'value': count, 'share': share})
    return items


def _recent_activity(placements, logs, scores):
    activity = []

    for placement in placements[:4]:
        organization = placement.organization.name if placement.organization else 'Unassigned organization'
        activity.append(
            {
                'type': 'placement',
                'title': f'Placement {placement.approval_status}',
                'description': f'{placement.student.get_full_name()} • {organization}',
                'status': placement.current_lifecycle_status,
                'time': placement.updated_at,
            }
        )

    for log in logs[:4]:
        organization = log.placement.organization.name if log.placement.organization else 'Unassigned organization'
        activity.append(
            {
                'type': 'log',
                'title': f'Week {log.week_number} log {log.review_status}',
                'description': f'{log.student.get_full_name()} • {organization}',
                'status': log.review_status,
                'time': log.updated_at,
            }
        )

    for score in scores[:4]:
        organization = score.placement.organization.name if score.placement.organization else 'Unassigned organization'
        activity.append(
            {
                'type': 'evaluation',
                'title': f'Grade {score.grade} released',
                'description': f'{score.student.get_full_name()} • {organization}',
                'status': score.grade,
                'time': score.computed_at,
            }
        )

    activity.sort(key=lambda item: item['time'] or timezone.now(), reverse=True)
    return [
        {**item, 'time': item['time'].strftime('%Y-%m-%d %H:%M') if item['time'] else 'Recently'}
        for item in activity[:8]
    ]


def _report_payload(date_from=None, date_to=None):
    placements_qs = _apply_date_window(
        Placement.objects.select_related('organization', 'student', 'approved_by', 'workplace_supervisor', 'academic_supervisor'),
        date_from,
        date_to,
    )
    logs_qs = _apply_date_window(
        WeeklyLog.objects.select_related('student', 'placement', 'placement__organization', 'reviewed_by'),
        date_from,
        date_to,
    )
    scores_qs = _apply_date_window(
        FinalInternshipScore.objects.select_related('student', 'placement', 'placement__organization', 'computed_by'),
        date_from,
        date_to,
        field_name='computed_at',
    )
    evaluations_qs = _apply_date_window(
        PlacementEvaluation.objects.select_related('student', 'placement', 'placement__organization', 'evaluated_by'),
        date_from,
        date_to,
    )

    users = User.objects.all()
    student_users = users.filter(role='student')
    supervisor_users = users.filter(role__in=['workplace_supervisor', 'academic_supervisor'])
    admin_users = users.filter(role='admin')

    pending_supervisors = supervisor_users.filter(admin_approved=False).count()
    placements = list(placements_qs)
    logs = list(logs_qs)
    scores = list(scores_qs)
    evaluations = list(evaluations_qs)

    approved_placements = [placement for placement in placements if placement.approval_status == Placement.APPROVAL_APPROVED]
    active_placements = [placement for placement in placements if placement.current_lifecycle_status == 'active']
    completed_placements = [placement for placement in placements if placement.current_lifecycle_status == 'completed']
    pending_placements = [placement for placement in placements if placement.approval_status == Placement.APPROVAL_PENDING]

    log_status_counts = Counter(log.review_status for log in logs)
    placement_status_counts = Counter(placement.approval_status for placement in placements)
    grade_counts = Counter(score.grade or 'N/A' for score in scores)
    organization_counts = Counter(
        placement.organization.name if placement.organization else 'Unassigned'
        for placement in placements
    )
    region_counts = Counter(
        placement.organization.region if placement.organization and placement.organization.region else 'Unassigned'
        for placement in placements
    )

    total_logs = len(logs)
    approved_logs = log_status_counts.get(WeeklyLog.REVIEW_APPROVED, 0)
    pending_reviews = log_status_counts.get(WeeklyLog.REVIEW_PENDING, 0) + log_status_counts.get(WeeklyLog.REVIEW_UNDER_REVIEW, 0)
    revision_logs = log_status_counts.get(WeeklyLog.REVIEW_NEEDS_REVISION, 0)
    late_logs = sum(1 for log in logs if log.is_late)

    final_scores_total = len(scores)
    average_final_score = _decimal_average([score.final_score for score in scores])
    evaluation_completion_rate = round((len(evaluations) / len(approved_placements)) * 100, 1) if approved_placements else 0

    grade_distribution = _top_items(grade_counts, limit=5, total=final_scores_total)
    organization_distribution = _top_items(organization_counts, limit=5, total=len(placements) or None)
    region_distribution = _top_items(region_counts, limit=5, total=len(placements) or None)

    top_performers = [
        {
            'student_name': score.student.get_full_name(),
            'organization': score.placement.organization.name if score.placement.organization else 'Unassigned',
            'final_score': str(score.final_score),
            'grade': score.grade,
            'remarks': score.remarks,
        }
        for score in sorted(scores, key=lambda item: item.final_score, reverse=True)[:5]
    ]

    alerts = []
    if pending_supervisors:
        alerts.append({
            'severity': 'warning',
            'title': 'Supervisor approvals pending',
            'count': pending_supervisors,
            'detail': 'Supervisor accounts still need administrative approval.',
        })
    if pending_placements:
        alerts.append({
            'severity': 'warning',
            'title': 'Placements awaiting review',
            'count': len(pending_placements),
            'detail': 'Submitted placements are still waiting for a decision.',
        })
    if pending_reviews:
        alerts.append({
            'severity': 'info',
            'title': 'Logs awaiting review',
            'count': pending_reviews,
            'detail': 'Weekly logs have not yet been actioned by supervisors.',
        })
    if late_logs:
        alerts.append({
            'severity': 'danger',
            'title': 'Late submissions detected',
            'count': late_logs,
            'detail': 'Some students submitted logs after the expected window.',
        })
    if placements and not scores:
        alerts.append({
            'severity': 'info',
            'title': 'No final scores released yet',
            'count': 0,
            'detail': 'Final internship scores have not been generated for the current data window.',
        })

    health_score = 100
    if pending_reviews:
        health_score -= min(25, pending_reviews * 2)
    if late_logs:
        health_score -= min(20, late_logs)
    if pending_supervisors:
        health_score -= min(20, pending_supervisors)
    if not scores:
        health_score -= 10
    health_score = max(0, health_score)

    stats = [
        { 'label': 'Students', 'value': student_users.count(), 'helper': 'Registered student accounts', 'accent': '#2E8B5B' },
        { 'label': 'Active Placements', 'value': len(active_placements), 'helper': 'Current internship placements', 'accent': '#5569E0' },
        { 'label': 'Pending Reviews', 'value': pending_reviews, 'helper': 'Logs still waiting for action', 'accent': '#F08C30' },
        { 'label': 'System Health', 'value': f'{health_score}%', 'helper': 'Operational visibility score', 'accent': '#0F7B5C' },
    ]

    return {
        'window': {
            'generated_at': timezone.now().isoformat(),
            'date_from': date_from.isoformat() if date_from else '',
            'date_to': date_to.isoformat() if date_to else '',
        },
        'stats': stats,
        'overview': {
            'users': {
                'students': student_users.count(),
                'supervisors': supervisor_users.count(),
                'admins': admin_users.count(),
                'pending_supervisor_approvals': pending_supervisors,
            },
            'placements': {
                'total': len(placements),
                'approved': len(approved_placements),
                'active': len(active_placements),
                'completed': len(completed_placements),
                'pending': len(pending_placements),
                'by_status': [
                    {'label': 'Approved', 'value': placement_status_counts.get(Placement.APPROVAL_APPROVED, 0)},
                    {'label': 'Pending', 'value': placement_status_counts.get(Placement.APPROVAL_PENDING, 0)},
                    {'label': 'Rejected', 'value': placement_status_counts.get(Placement.APPROVAL_REJECTED, 0)},
                    {'label': 'Cancelled', 'value': placement_status_counts.get(Placement.APPROVAL_CANCELLED, 0)},
                ],
                'by_organization': organization_distribution,
                'by_region': region_distribution,
            },
            'logbook': {
                'total': total_logs,
                'approved': approved_logs,
                'pending_review': pending_reviews,
                'revisions': revision_logs,
                'late_submissions': late_logs,
                'approval_rate': round((approved_logs / total_logs) * 100, 1) if total_logs else 0,
            },
            'performance': {
                'average_final_score': str(average_final_score),
                'evaluation_completion_rate': evaluation_completion_rate,
                'final_scores': final_scores_total,
                'grade_distribution': grade_distribution,
                'top_performers': top_performers,
            },
        },
        'alerts': alerts,
        'recent_activity': _recent_activity(placements, logs, scores),
    }


def _render_csv(report):
    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(['Metric', 'Value', 'Context'])
    for item in report['stats']:
        writer.writerow([item['label'], item['value'], item.get('helper', '')])

    writer.writerow([])
    writer.writerow(['Placement Status', 'Count', 'Share'])
    for item in report['overview']['placements']['by_status']:
        writer.writerow([item['label'], item['value'], ''])

    writer.writerow([])
    writer.writerow(['Top Organization', 'Placements', 'Share'])
    for item in report['overview']['placements']['by_organization']:
        writer.writerow([item['label'], item['value'], f"{item['share']}%"])

    writer.writerow([])
    writer.writerow(['Grade', 'Count', 'Share'])
    for item in report['overview']['performance']['grade_distribution']:
        writer.writerow([item['label'], item['value'], f"{item['share']}%"])

    writer.writerow([])
    writer.writerow(['Alert', 'Count', 'Detail'])
    for item in report['alerts']:
        writer.writerow([item['title'], item['count'], item['detail']])

    return output.getvalue()


def _render_xlsx(report):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = 'Summary'

    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)

    def write_block(sheet, title, rows, start_row=1):
        row = start_row
        sheet.cell(row=row, column=1, value=title).font = title_font
        row += 2
        for index, header in enumerate(rows[0], start=1):
            cell = sheet.cell(row=row, column=index, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        row += 1
        for data_row in rows[1:]:
            for index, value in enumerate(data_row, start=1):
                sheet.cell(row=row, column=index, value=value)
            row += 1
        return row + 1

    summary_rows = [['Metric', 'Value', 'Context']]
    summary_rows.extend([[item['label'], item['value'], item.get('helper', '')] for item in report['stats']])
    row = write_block(summary_sheet, 'Executive Summary', summary_rows, 1)

    summary_sheet.cell(row=row, column=1, value='Alerts').font = title_font
    row += 2
    for index, header in enumerate(['Alert', 'Count', 'Detail'], start=1):
        summary_sheet.cell(row=row, column=index, value=header).font = header_font
    row += 1
    for alert in report['alerts']:
        summary_sheet.cell(row=row, column=1, value=alert['title'])
        summary_sheet.cell(row=row, column=2, value=alert['count'])
        summary_sheet.cell(row=row, column=3, value=alert['detail'])
        row += 1

    placements_sheet = workbook.create_sheet('Placements')
    placement_rows = [['Category', 'Count', 'Share']]
    placement_rows.extend([[item['label'], item['value'], item['share']] for item in report['overview']['placements']['by_status']])
    write_block(placements_sheet, 'Placement Status', placement_rows, 1)

    placements_sheet.cell(row=10, column=1, value='Organizations').font = title_font
    placement_org_headers = ['Organization', 'Placements', 'Share']
    for index, header in enumerate(placement_org_headers, start=1):
        placements_sheet.cell(row=12, column=index, value=header).font = header_font
    placement_row = 13
    for item in report['overview']['placements']['by_organization']:
        placements_sheet.cell(row=placement_row, column=1, value=item['label'])
        placements_sheet.cell(row=placement_row, column=2, value=item['value'])
        placements_sheet.cell(row=placement_row, column=3, value=item['share'])
        placement_row += 1

    logbook_sheet = workbook.create_sheet('Logbook')
    logbook_rows = [['Metric', 'Count']]
    logbook_rows.extend([
        ['Total Logs', report['overview']['logbook']['total']],
        ['Approved', report['overview']['logbook']['approved']],
        ['Pending Review', report['overview']['logbook']['pending_review']],
        ['Revisions', report['overview']['logbook']['revisions']],
        ['Late Submissions', report['overview']['logbook']['late_submissions']],
        ['Approval Rate', report['overview']['logbook']['approval_rate']],
    ])
    write_block(logbook_sheet, 'Logbook Overview', logbook_rows, 1)

    performance_sheet = workbook.create_sheet('Performance')
    performance_rows = [['Grade', 'Count', 'Share']]
    performance_rows.extend([[item['label'], item['value'], item['share']] for item in report['overview']['performance']['grade_distribution']])
    write_block(performance_sheet, 'Grade Distribution', performance_rows, 1)

    performers_start = 10
    performance_sheet.cell(row=performers_start, column=1, value='Top Performers').font = title_font
    for index, header in enumerate(['Student', 'Organization', 'Score', 'Grade', 'Remarks'], start=1):
        performance_sheet.cell(row=performers_start + 2, column=index, value=header).font = header_font
    performer_row = performers_start + 3
    for item in report['overview']['performance']['top_performers']:
        performance_sheet.cell(row=performer_row, column=1, value=item['student_name'])
        performance_sheet.cell(row=performer_row, column=2, value=item['organization'])
        performance_sheet.cell(row=performer_row, column=3, value=item['final_score'])
        performance_sheet.cell(row=performer_row, column=4, value=item['grade'])
        performance_sheet.cell(row=performer_row, column=5, value=item['remarks'])
        performer_row += 1

    for sheet in workbook.worksheets:
        sheet.freeze_panes = 'A3'
        for column_cells in sheet.columns:
          max_length = 0
          column = column_cells[0].column_letter
          for cell in column_cells:
              if cell.value is None:
                  continue
              max_length = max(max_length, len(str(cell.value)))
          sheet.column_dimensions[column].width = min(max_length + 4, 36)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _render_pdf(report):
    output = BytesIO()
    document_title = 'AILES Insights Report'
    page_width, page_height = landscape(letter)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='SmallBody', parent=styles['BodyText'], fontSize=8.5, leading=10.5, spaceAfter=4))

    story = []
    story.append(Paragraph(document_title, styles['Title']))
    story.append(Paragraph(f"Generated at {report['window']['generated_at']}", styles['SmallBody']))
    if report['window']['date_from'] or report['window']['date_to']:
        story.append(Paragraph(
            f"Reporting window: {report['window']['date_from'] or 'Start'} to {report['window']['date_to'] or 'Now'}",
            styles['SmallBody'],
        ))
    story.append(Spacer(1, 0.2 * inch))

    summary_rows = [['Metric', 'Value', 'Context']]
    summary_rows.extend([[item['label'], str(item['value']), item.get('helper', '')] for item in report['stats']])
    summary_table = Table(summary_rows, colWidths=[2.1 * inch, 1.0 * inch, 4.9 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A7A57')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('LEADING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D0D7E2')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#F7FAFC')]),
    ]))
    story.append(Paragraph('Executive Summary', styles['Heading2']))
    story.append(summary_table)
    story.append(Spacer(1, 0.2 * inch))

    placement_rows = [['Status', 'Count', 'Share']]
    placement_rows.extend([[item['label'], str(item['value']), f"{item['share']}%"] for item in report['overview']['placements']['by_status']])
    placement_table = Table(placement_rows, colWidths=[2.0 * inch, 1.1 * inch, 1.1 * inch])
    placement_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E8B5B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D0D7E2')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#F7FAFC')]),
    ]))

    grade_rows = [['Grade', 'Count', 'Share']]
    grade_rows.extend([[item['label'], str(item['value']), f"{item['share']}%"] for item in report['overview']['performance']['grade_distribution']])
    grade_table = Table(grade_rows, colWidths=[2.0 * inch, 1.1 * inch, 1.1 * inch])
    grade_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F08C30')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D0D7E2')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#F7FAFC')]),
    ]))

    two_col = Table([[placement_table, grade_table]], colWidths=[3.5 * inch, 3.5 * inch])
    two_col.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    story.append(Paragraph('Snapshot Charts', styles['Heading2']))
    story.append(two_col)
    story.append(Spacer(1, 0.2 * inch))

    alerts_rows = [['Alert', 'Count', 'Detail']]
    alerts_rows.extend([[item['title'], str(item['count']), item['detail']] for item in report['alerts']])
    alerts_table = Table(alerts_rows, colWidths=[2.4 * inch, 0.8 * inch, 4.3 * inch])
    alerts_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C0392B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D0D7E2')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#FFF7F7')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(Paragraph('System Alerts', styles['Heading2']))
    story.append(alerts_table)

    def _canvas(canvas, doc):
        canvas.setTitle(document_title)
        canvas.setAuthor('AILES')
        canvas.setFont('Helvetica', 8)
        canvas.drawRightString(page_width - 0.4 * inch, 0.35 * inch, f'Page {doc.page}')

    from reportlab.platypus import SimpleDocTemplate

    doc = SimpleDocTemplate(output, pagesize=landscape(letter), rightMargin=0.45 * inch, leftMargin=0.45 * inch, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    doc.build(story, onFirstPage=_canvas, onLaterPages=_canvas)
    return output.getvalue()


def _export_report(report, export_format='csv'):
    normalized = (export_format or 'csv').lower()

    if normalized == 'xlsx':
        content = _render_xlsx(report)
        response = HttpResponse(content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="insights-report.xlsx"'
        return response

    if normalized == 'pdf':
        content = _render_pdf(report)
        response = HttpResponse(content, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="insights-report.pdf"'
        return response

    content = _render_csv(report)
    response = HttpResponse(content, content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="insights-report.csv"'
    return response


class InsightDashboardView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsAdmin]

    def get(self, request):
        payload = _report_payload()
        return Response(payload)


class AdminReportSummaryView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsAdmin]

    def get(self, request):
        date_from = _parse_date(request.query_params.get('date_from'))
        date_to = _parse_date(request.query_params.get('date_to'))
        report = _report_payload(date_from=date_from, date_to=date_to)
        return Response(report)


class AdminReportExportView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsAdmin]

    def get(self, request):
        date_from = _parse_date(request.query_params.get('date_from'))
        date_to = _parse_date(request.query_params.get('date_to'))
        export_format = request.query_params.get('format', 'csv')
        report = _report_payload(date_from=date_from, date_to=date_to)
        return _export_report(report, export_format=export_format)